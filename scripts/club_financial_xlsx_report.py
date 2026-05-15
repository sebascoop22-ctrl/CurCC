"""
Summarise Club financial tracking data.xlsx (structure + any data rows).

Usage:
  python scripts/club_financial_xlsx_report.py [path/to/file.xlsx]

Conflicts vs current Cooper DB (high level):
- financial_rules: DB has department, venue, male/female/base rates, bonus_*, logic_type.
  Excel adds guestlist/table subdivisions, min age, bar spend, deposits — stored in sheet_extension JSON.
- financial_promoters: DB has commission %, contact, notes. Excel adds payment schedule, bank, tax —
  stored in sheet_extension; consider also financial_payees for canonical bank payloads.
- Club payment info / nightlife job / club job sheets: no first-class tables yet; use financial_payees
  (payment_details jsonb) and financial_bookings / promoter_jobs as appropriate, or extend sheet_extension
  on clubs (not added here — ask before new clubs columns).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("pip install openpyxl") from e


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "Club financial tracking data.xlsx"
    if not path.is_file():
        raise SystemExit(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, object] = {"file": str(path), "sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[object | None]] = []
        nonempty = 0
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            if any(c is not None and str(c).strip() != "" for c in cells):
                nonempty += 1
            if len(rows) < 80:
                rows.append(cells)
        out["sheets"][name] = {
            "nonempty_row_estimate": nonempty,
            "first_rows_preview": rows[:40],
        }
    wb.close()
    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
